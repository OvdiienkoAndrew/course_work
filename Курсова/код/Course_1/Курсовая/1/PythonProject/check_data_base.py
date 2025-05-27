from tkinter import messagebox, filedialog

import pandas as pd

from resources import loading_window
import sqlite3
import openpyxl

def get_cell_str(df, i, j):
    j-=1
    i-=1
    if df.iloc[i, j] is None or pd.isna(df.iloc[i, j]):
        return ""
    return str(df.iloc[i, j])

class Person:
    def __init__(self, name, surname, patronymic, position,counter):
        self.name = name
        self.surname = surname
        self.patronymic = patronymic
        self.position = position
        self.counter = counter
        self.bid = 0
        self.result = 0
        self.min = 0
        self.max = 0
        self.MING = 0
        self.MAXG = 0
        self.total_year = 0

def read_numbers(df,people, data, k,name_db):
     try:
            for i in range(2,1048576):
                position = ""
                helper = get_cell_str(df, i, 2).lower().replace(' ', "")
                if helper[0] == 'з':
                    position = "заф. кафедри"
                elif helper[0] == 'п':
                    position = "професор"
                elif helper[0] == 'д':
                    position = "доцент"
                elif helper[0] == 'с':
                    position = "ст. викладач"
                elif helper[0] == 'а':
                    position = "асистент"
                elif helper[0] == 'в':
                    if helper[1] == '.':
                        position = "в.о. заф. кафедри"
                    else:
                        position = "викладач"

                helper = get_cell_str(df, i, 1)
                name = ""
                surname = ""
                patronymic = ""
                parts = helper.split()
                if len(parts) >= 1:
                    surname = parts[0]
                    if len(parts) >= 2:
                        name = parts[1]
                    if len(parts) >= 3:
                        patronymic = parts[2]

                was = False
                for person in people:
                    if surname == person.surname and person.name == name and person.patronymic == patronymic:
                        person.counter += data[k]
                        was = True
                        break

                if was is True:
                    continue

                conn = sqlite3.connect(name_db)

                cursor = conn.cursor()

                person_info = cursor.execute("SELECT * FROM ЛЮДСЬКА_ІНФОРМАЦІЯ").fetchall()
                academic_year_info = cursor.execute("SELECT * FROM АКАДЕМІЧНИЙ_РІК").fetchall()
                conn.close()

                was = False
                j = -1
                for person_db in person_info:
                    j += 1

                    if str(person_db[1]) == str(name) and str(person_db[2]) == surname and str(
                            person_db[3]) == patronymic and len(str(academic_year_info[j][2])) <= 0:
                        was = True
                        break

                if was is False:
                    continue

                people.append(Person(name, surname, patronymic, position, data[k]))
     except Exception:
       i=1

def check_button_click(root,name_db):
    file = filedialog.askopenfilename(
        title="Оберіть Excel файл",
        filetypes=[("Excel Files", "*.xlsx")]
    )

    if file == "":
        messagebox.showinfo("Помилка", "Файл не обрано або має валідну помилку в назві!")
        return

    loading = loading_window(root)


    conn = sqlite3.connect(name_db)

    cursor = conn.cursor()

    job_info = cursor.execute("SELECT * FROM РОБОЧА_ІНФОРМАЦІЯ").fetchall()
    person_info = cursor.execute("SELECT * FROM ЛЮДСЬКА_ІНФОРМАЦІЯ").fetchall()
    vacancy_info = cursor.execute("SELECT * FROM ІНФОРМАЦІЯ_ВАКАНСІЯ").fetchall()
    first_semester_info = cursor.execute("SELECT * FROM ПЕРШИЙ_СЕМЕСТР").fetchall()
    second_semester_info = cursor.execute("SELECT * FROM ДРУГИЙ_СЕМЕСТР").fetchall()
    academic_year_info = cursor.execute("SELECT * FROM АКАДЕМІЧНИЙ_РІК").fetchall()
    code_and_year_info = cursor.execute("SELECT * FROM КОД_РІК").fetchall()

    conn.close()

    wb = openpyxl.load_workbook(file)

    if 'Permission' in wb.sheetnames:
        del wb['Permission']
    if 'Load' in wb.sheetnames:
        del wb['Load']

    sheetname_1 = "1"
    sheetname_2 = "2"
    sheetname_3 = "3"
    sheetname_4 = "4"
    sheetname_5 = "5"
    sheetname_6 = "6"
    sheetname_7 = "7"
    sheetname_8 = "8"

    if sheetname_1 in wb.sheetnames:
        i=1
    else:
        wb.create_sheet(sheetname_1)

    if sheetname_2 in wb.sheetnames:
        i = 1
    else:
        wb.create_sheet(sheetname_2)

    if sheetname_3 in wb.sheetnames:
        i=1
    else:
        wb.create_sheet(sheetname_3)

    if sheetname_4 in wb.sheetnames:
        i = 1
    else:
        wb.create_sheet(sheetname_4)

    if sheetname_5 in wb.sheetnames:
        i = 1
    else:
        wb.create_sheet(sheetname_5)

    if sheetname_6 in wb.sheetnames:
        i = 1
    else:
        wb.create_sheet(sheetname_6)

    if sheetname_7 in wb.sheetnames:
        i = 1
    else:
        wb.create_sheet(sheetname_7)

    if sheetname_8 in wb.sheetnames:
        i = 1
    else:
        wb.create_sheet(sheetname_8)

    permission_sheet = wb.create_sheet('Permission')
    load_sheet = wb.create_sheet('Load')

    try:
        wb.save(file)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", "Файл відкритий!\n\nЗакрийте файл та повторіть спробу!")
        return



    index_assistant = 1

    with sqlite3.connect(name_db) as db:
        cursor = db.cursor()

        сheck_info_as_st = """
                INSERT INTO ПЕРЕВІРКА_АСИСТЕНТ_СТ_ВИКЛАДАЧ ("прізвище","ім'я","по-батькові", "кафедра","роки","помилка")
                VALUES (?, ?,?,?,?,?)
                """
        delete_query = """
                        DELETE FROM ПЕРЕВІРКА_АСИСТЕНТ_СТ_ВИКЛАДАЧ
                        WHERE "прізвище" = ? AND "ім'я" = ? AND "по-батькові" = ? AND  "кафедра" = ? AND "роки" = ?
                        """

        try:
            for i in range(0, len(person_info)):
                if str(person_info[i][1]) != "":
                    if str(job_info[i][1]) == "асистент" or str(job_info[i][1]) == "ст. викладач":
                        for j in range(0,4):
                            try:
                                Full_code = "(" + str(code_and_year_info[i][1]) + ")"
                                values = (str(person_info[i][2]), str(person_info[i][1]), str(person_info[i][3]), Full_code,str(code_and_year_info[i][2]))

                                cursor.execute(delete_query, values)
                                db.commit()
                            except Exception:
                                continue
        except Exception:
            index_assistant=1

        index_assistant=1

        try:
            for i in range(0, len(person_info)):
                if str(person_info[i][1]) != "":
                    if str(job_info[i][1]) == "асистент":
                        if str(academic_year_info[i][3]) != "0.0" or str(academic_year_info[i][6]) != "0.0" != 0.0 or str(academic_year_info[i][7]) != "0.0" != 0.0 or str(academic_year_info[i][9]) != "0.0":
                            Full_name = str(person_info[i][2]) + " " + str(person_info[i][1]) + " " + str(person_info[i][3])
                            permission_sheet.cell(row=index_assistant, column=1, value=Full_name)
                            Full_code = "(" + str(code_and_year_info[i][1]) + ")"
                            permission_sheet.cell(row=index_assistant, column=2, value=Full_code)
                            Full_year = str(code_and_year_info[i][2])
                            permission_sheet.cell(row=index_assistant, column=3, value=Full_year)
                            index_assistant+=1

                            if str(academic_year_info[i][3]) != "0.0":
                                temp = "Асистент веде лекції в розмірі " + str(academic_year_info[i][3]) + " годин."
                                permission_sheet.cell(row=index_assistant, column=4, value=temp)
                                index_assistant += 1
                                cursor.execute(сheck_info_as_st, (str(person_info[i][2]), str(person_info[i][1]), str(person_info[i][3]), Full_code, Full_year, temp))
                                db.commit()

                            if str(academic_year_info[i][7]) != "0.0":
                                temp = "Асистент веде консультації перед екзаменом в розмірі " + str(academic_year_info[i][7]) + " годин."
                                permission_sheet.cell(row=index_assistant, column=4, value=temp)
                                index_assistant += 1
                                cursor.execute(сheck_info_as_st, (
                                str(person_info[i][2]), str(person_info[i][1]), str(person_info[i][3]), Full_code,
                                Full_year, temp))
                                db.commit()

                            if str(academic_year_info[i][6]) != "0.0":
                                temp = "Асистент веде екзамен в розмірі " + str(academic_year_info[i][6]) + " годин."
                                permission_sheet.cell(row=index_assistant, column=4, value=temp)
                                index_assistant += 1
                                cursor.execute(сheck_info_as_st, (
                                str(person_info[i][2]), str(person_info[i][1]), str(person_info[i][3]), Full_code,
                                Full_year, temp))
                                db.commit()

                            if str(academic_year_info[i][9]) != "0.0":
                                temp = "Асистент веде кваліфікаційну роботу в розмірі " + str(academic_year_info[i][9]) + " годин."
                                permission_sheet.cell(row=index_assistant, column=4, value=temp)
                                index_assistant += 1
                                cursor.execute(сheck_info_as_st, (
                                str(person_info[i][2]), str(person_info[i][1]), str(person_info[i][3]), Full_code,
                                Full_year, temp))
                                db.commit()

                    if str(job_info[i][1]) == "ст. викладач":
                        work = True
                        if str(academic_year_info[i][9]) != "0.0":
                            work = False
                        qualification = True
                        if str(academic_year_info[i][3]) != "0.0":
                            loading.destroy()
                            result = messagebox.askyesno("Попередження", "Вчена рада надала дозвіл вести лекції ст. викладачу " + str(person_info[i][2]) + " " + str(person_info[i][1]) + " " + str(person_info[i][3]) + "(" + str(code_and_year_info[i][1]) + ") у " + str(code_and_year_info[i][2]) + " роках?")
                            loading = loading_window(root)
                            if result:
                                qualification = True
                            else:
                                qualification = False

                        if work == False or qualification == False:

                            Full_name = str(person_info[i][2]) + " " + str(person_info[i][1]) + " " + str(person_info[i][3])
                            permission_sheet.cell(row=index_assistant, column=1, value=Full_name)
                            Full_code = "(" + str(code_and_year_info[i][1]) + ")"
                            permission_sheet.cell(row=index_assistant, column=2, value=Full_code)
                            Full_year = str(code_and_year_info[i][2])
                            permission_sheet.cell(row=index_assistant, column=3, value=Full_year)
                            index_assistant += 1

                            if work == False:
                                temp = "cт. викладач веде кваліфікаційну роботу в розмірі " + str(academic_year_info[i][9]) + " годин."
                                permission_sheet.cell(row=index_assistant, column=4, value=temp)
                                index_assistant += 1
                                cursor.execute(сheck_info_as_st, (str(person_info[i][2]), str(person_info[i][1]), str(person_info[i][3]), Full_code,Full_year, temp))
                                db.commit()

                            if qualification == False:
                                temp = "cт. викладач веде лекції без дозволу вченої ради в розмірі " + str(academic_year_info[i][3]) + " годин."
                                permission_sheet.cell(row=index_assistant, column=4, value=temp)
                                index_assistant += 1
                                cursor.execute(сheck_info_as_st, (
                                str(person_info[i][2]), str(person_info[i][1]), str(person_info[i][3]), Full_code,
                                Full_year, temp))
                                db.commit()

        except Exception:
            j=1

    datas=[]

    with open("resources/settings/file.txt", "r", encoding="utf-8") as filename:
        datas = [list(map(float, line.split())) for line in filename.readlines()]

    data = datas[0]



    try:
        df = pd.read_excel(file, sheet_name=sheetname_1, engine="openpyxl", header=None)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", f"Лист: \"{sheetname_1}\" - не знайдено!")
        return

    people = []

    try:
        for i in range(2,1048576):
            position = ""
            helper = get_cell_str(df, i, 2).lower().replace(' ', "")
            if helper[0] == 'з':
                position = "заф. кафедри"
            elif helper[0] == 'п':
                position = "професор"
            elif helper[0] == 'д':
                position = "доцент"
            elif helper[0] == 'с':
                position = "ст. викладач"
            elif helper[0] == 'а':
                position = "асистент"
            elif helper[0] == 'в':
                if helper[1] == '.':
                    position = "в.о. заф. кафедри"
                else:
                    position = "викладач"

            helper = str(get_cell_str(df, i, 1)).replace(' ',',').replace(',,',',').replace(',',' ')
            name = ""
            surname = ""
            patronymic = ""
            parts = helper.split()
            if len(parts) >= 1:
                surname = str(parts[0])
                if len(parts) >= 2:
                    name = str(parts[1])
                if len(parts) >= 3:
                    patronymic = str(parts[2]).replace(' ', '')

            was=False
            for person in people:
                if surname== person.surname and person.name==name and person.patronymic==patronymic:
                    person.counter += data[4]
                    was=True
                    break

            if was is True:
                continue

            conn = sqlite3.connect(name_db)

            cursor = conn.cursor()

            person_info = cursor.execute("SELECT * FROM ЛЮДСЬКА_ІНФОРМАЦІЯ").fetchall()
            academic_year_info = cursor.execute("SELECT * FROM АКАДЕМІЧНИЙ_РІК").fetchall()
            conn.close()

            was = False
            j = -1
            for person_db in person_info:
                j+=1

                if str(person_db[1]) == str(name) and str(person_db[2]) == surname and str(person_db[3]) == patronymic and len(str(academic_year_info[j][2])) <= 0:
                    was = True
                    break

            if was is False:
                continue

            people.append(Person(name, surname, patronymic, position, data[4]))
    except Exception:
       i=1


    try:
        df = pd.read_excel(file, sheet_name=sheetname_2, engine="openpyxl", header=None)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", f"Лист: \"{sheetname_2}\" - не знайдено!")


    read_numbers(df, people, data, 5,name_db)

    try:
        df = pd.read_excel(file, sheet_name=sheetname_3, engine="openpyxl", header=None)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", f"Лист: \"{sheetname_3}\" - не знайдено!")


    read_numbers(df, people, data, 6,name_db)

    try:
        df = pd.read_excel(file, sheet_name=sheetname_4, engine="openpyxl", header=None)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", f"Лист: \"{sheetname_4}\" - не знайдено!")


    read_numbers(df, people, data, 7,name_db)

    try:
        df = pd.read_excel(file, sheet_name=sheetname_5, engine="openpyxl", header=None)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", f"Лист: \"{sheetname_5}\" - не знайдено!")


    read_numbers(df, people, data, 8,name_db)

    try:
        df = pd.read_excel(file, sheet_name=sheetname_6, engine="openpyxl", header=None)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", f"Лист: \"{sheetname_6}\" - не знайдено!")


    read_numbers(df, people, data, 9,name_db)

    try:
        df = pd.read_excel(file, sheet_name=sheetname_7, engine="openpyxl", header=None)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", f"Лист: \"{sheetname_7}\" - не знайдено!")


    read_numbers(df, people, data, 10,name_db)




    conn = sqlite3.connect(name_db)

    cursor = conn.cursor()

    job_info = cursor.execute("SELECT * FROM РОБОЧА_ІНФОРМАЦІЯ").fetchall()
    person_info = cursor.execute("SELECT * FROM ЛЮДСЬКА_ІНФОРМАЦІЯ").fetchall()
    vacancy_info = cursor.execute("SELECT * FROM ІНФОРМАЦІЯ_ВАКАНСІЯ").fetchall()
    first_semester_info = cursor.execute("SELECT * FROM ПЕРШИЙ_СЕМЕСТР").fetchall()
    second_semester_info = cursor.execute("SELECT * FROM ДРУГИЙ_СЕМЕСТР").fetchall()
    academic_year_info = cursor.execute("SELECT * FROM АКАДЕМІЧНИЙ_РІК").fetchall()
    code_and_year_info = cursor.execute("SELECT * FROM КОД_РІК").fetchall()

    conn.close()

    i=0
    j=0


    for value_person_info in person_info:
        for person in people:
            if person.name == str(value_person_info[1]) and person.surname == str(value_person_info[2]) and person.patronymic == str(value_person_info[3]) and len(str(academic_year_info[i][2])) <=0:
                 person.bid = float(str(academic_year_info[i][1]))
                 person.total_year = float(str(academic_year_info[i][22]))
            j+=1
        i+=1


    conn = sqlite3.connect(name_db)

    cursor = conn.cursor()

    job_info = cursor.execute("SELECT * FROM РОБОЧА_ІНФОРМАЦІЯ").fetchall()
    person_info = cursor.execute("SELECT * FROM ЛЮДСЬКА_ІНФОРМАЦІЯ").fetchall()
    vacancy_info = cursor.execute("SELECT * FROM ІНФОРМАЦІЯ_ВАКАНСІЯ").fetchall()
    first_semester_info = cursor.execute("SELECT * FROM ПЕРШИЙ_СЕМЕСТР").fetchall()
    second_semester_info = cursor.execute("SELECT * FROM ДРУГИЙ_СЕМЕСТР").fetchall()
    academic_year_info = cursor.execute("SELECT * FROM АКАДЕМІЧНИЙ_РІК").fetchall()
    code_and_year_info = cursor.execute("SELECT * FROM КОД_РІК").fetchall()

    conn.close()


    i = -1
    for person in person_info:
        i += 1
        for person2 in people:

            if str(person[1]) == str(person2.name) and str(person[2]) == str(person2.surname) and str(person[3]) == str(
                    person2.patronymic) and str(academic_year_info[i][2]) == "":
                job_info.pop(i)
                person_info.pop(i)
                vacancy_info.pop(i)
                first_semester_info.pop(i)
                second_semester_info.pop(i)
                academic_year_info.pop(i)
                code_and_year_info.pop(i)

    k = 2

    with sqlite3.connect(name_db) as db:
        cursor = db.cursor()

        add_from_file = """
         INSERT INTO ПЕРЕВІРКА ("прізвище","ім'я","по-батькові", "ставка","знак_сумісності","загальна_кількость_годин","мінімум","максимум", "загальний_мінімум","загальний_максимум")
         VALUES (?,?,?,?,?,?,?,?,?,?)
         """

        delete_from_file = """
                        DELETE FROM ПЕРЕВІРКА
                        WHERE "прізвище" = ? AND "ім'я" = ? AND "по-батькові" = ? AND  "ставка"=? AND "знак_сумісності" =?
                        """




        i=2

        l=-1
        for person in people:


            try:
                values = (person.surname, person.name, person.patronymic, person.bid, "")

                cursor.execute(delete_from_file, values)
                db.commit()

            except Exception:
                q=0

            person.result = round(data[0]*person.bid-person.counter)
            person.min = round(person.bid*data[3])
            person.max = round(person.min*(100+data[1]))

            person.MING = round(data[3]*person.bid)
            person.MAXG = round(data[2]*person.bid)

            if person.result < person.min:
                person.result = person.min
            if person.result > person.max:
                person.result = person.max

            load_sheet.cell(row=1, column=1, value="ПІБ, річна ставка, річні години")
            load_sheet.cell(row=1, column=2, value="Мінімум (год)")
            load_sheet.cell(row=1, column=3, value="Максимум (год)")

            load_sheet.cell(row=1, column=4, value="Загальний мінімум (год)")
            load_sheet.cell(row=1, column=5, value="Загальний максимум (год)")

            l+=1
            person.total_year = float(academic_year_info[l][22])
            if person.total_year < person.MING or person.total_year > person.MAXG or person.total_year < person.min or person.total_year > person.max:


                loading.destroy()
                temp = f"Мінімальна кількість годин: {person.min}.\nМаксимальна кількість годин: {person.max}.\nЗагальний мінімум годин: {person.MING}.\nЗагальний максимум годин: {person.MAXG}.\nДопустимо для {person.surname} {person.name} {person.patronymic}?"
                result = messagebox.askyesno("Попередження",temp)
                loading = loading_window(root)
                if result:
                    continue
                else:

                    SNP = person.surname + " " + person.name + " " + person.patronymic + " ставка: " + str(person.bid) + ", загальна кількість годин: " + str(person.total_year)
                    load_sheet.cell(row=i, column=1, value=SNP)
                    load_sheet.cell(row=i, column=2, value=str(person.min))
                    load_sheet.cell(row=i, column=3, value=str(person.max))
                    load_sheet.cell(row=i, column=4, value=str(person.MING))
                    load_sheet.cell(row=i, column=5, value=str(person.MAXG))



                    cursor.execute(add_from_file, (person.surname, person.name, person.patronymic,person.bid,"",person.total_year,person.min,person.max,person.MING,person.MAXG))
                    db.commit()
                    i+=1
                    k=i


    try:
        wb.save(file)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", "Файл відкритий!\n\nЗакрийте файл та повторіть спробу!")


    with sqlite3.connect(name_db) as db:
        cursor = db.cursor()

        add_from_file = """
            INSERT INTO ПЕРЕВІРКА ("прізвище","ім'я","по-батькові", "ставка","знак_сумісності","загальна_кількость_годин","мінімум","максимум", "загальний_мінімум","загальний_максимум")
            VALUES (?,?,?,?,?,?,?,?,?,?)
            """

        delete_from_file = """
                                DELETE FROM ПЕРЕВІРКА
                                WHERE "прізвище" = ? AND "ім'я" = ? AND "по-батькові" = ? AND  "ставка"=? AND "знак_сумісності" =?
                                """

        i=k

        load_sheet.cell(row=1, column=1, value="ПІБ (Вакансія, номер), річна ставка, річні години")
        load_sheet.cell(row=1, column=2, value="Мінімум (год)")
        load_sheet.cell(row=1, column=3, value="Максимум (год)")

        load_sheet.cell(row=1, column=4, value="Загальний мінімум (год)")
        load_sheet.cell(row=1, column=5, value="Загальний максимум (год)")

        k=-1
        for person in person_info:



            k+=1
            if len(str(person_info[k][1])) <= 0:
                surname = ""
                name = str(vacancy_info[k][1])
                patronymic = str(vacancy_info[k][2])
            else:
                surname = str(person_info[k][1])
                name = str(person_info[k][2])
                patronymic = str(person_info[k][3])


            try:
                values = (surname, name, patronymic, str(academic_year_info[k][1]), str(academic_year_info[k][2]))

                cursor.execute(delete_from_file, values)
                db.commit()
            except Exception:
                q = 0


            MING = round(float(data[3]) * float(academic_year_info[k][1]))
            MAXG = round(float(data[2]) *float(academic_year_info[k][1]))
            total = round(float(academic_year_info[i][22]))

            if MING <= total and total <= MAXG:
                continue

            loading.destroy()
            temp = f"Загальний мінімум годин: {MING}.\nЗагальний максимум годин: {MAXG}.\nЗагальна кількість годин за рік: {total}\nДопустимо для {surname} {name} {patronymic}?"
            result = messagebox.askyesno("Попередження", temp)
            loading = loading_window(root)

            if result:
                continue

            SNP =""
            if len(str(person_info[k][1]))<=0:
                surname = ""
                name = str(vacancy_info[k][1])
                patronymic = str(vacancy_info[k][2])
                SNP = str(vacancy_info[k][1])+" "+str(vacancy_info[k][2])+", "
            else:
                surname = str(person_info[k][1])
                name = str(person_info[k][2])
                patronymic = str(person_info[k][3])
                SNP = str(person_info[k][2])+" "+str(person_info[k][1])+" "+str(person_info[k][3])+", "

            SNP+= "cтавка: "+str(academic_year_info[k][1])+", "
            if len(str(academic_year_info[k][2]))>0:
                SNP += str(academic_year_info[k][2]) + ", "
            SNP += "загальна кількость годин: "+str(total)

            load_sheet.cell(row=i, column=1, value=SNP)
            load_sheet.cell(row=i, column=2, value=0)
            load_sheet.cell(row=i, column=3, value=0)
            load_sheet.cell(row=i, column=4, value=MING)
            load_sheet.cell(row=i, column=5, value=MAXG)

            cursor.execute(add_from_file, (surname, name, patronymic, str(academic_year_info[k][1]), str(academic_year_info[k][2]), total, 0, 0,MING, MAXG))
            db.commit()

            i+=1

    try:
        wb.save(file)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", "Файл відкритий!\n\nЗакрийте файл та повторіть спробу!")


    loading.destroy()


