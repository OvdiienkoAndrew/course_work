from tkinter import messagebox, filedialog
from resources import loading_window
import sqlite3
import openpyxl


def print_db_with_person(sheet_StudyLoadPerson):
    sheet_StudyLoadPerson.cell(row=1, column=1, value="Назва кафедри")
    sheet_StudyLoadPerson.cell(row=1, column=2, value="Ім'я")
    sheet_StudyLoadPerson.cell(row=1, column=3, value="Прізвище")
    sheet_StudyLoadPerson.cell(row=1, column=4, value="По-батькові")
    sheet_StudyLoadPerson.cell(row=1, column=5, value="Посада")
    sheet_StudyLoadPerson.cell(row=1, column=6, value="Вчене звання, вчена ступінь")
    sheet_StudyLoadPerson.cell(row=1, column=7, value="Роки")

    sheet_StudyLoadPerson.cell(row=1, column=8, value="Ставка (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=9, value="Знак сумісності (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=10, value="Лекції (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=11, value="Практичні (семінарські) заняття (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=12, value="Лабораторні роботи (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=13, value="Екзамени (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=14, value="Консультації перед екзаменами (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=15, value="Заліки (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=16, value="Кваліфікаційні робота (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=17, value="Атестаційні екзамени (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=18, value="Виробнича практика (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=19, value="Навчальна практика (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=20, value="Поточні консультації (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=21, value="Індивідуальні (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=22, value="Курсові роботи (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=23, value="Аспірантські екзамени (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=24, value="Керівництво аспірантами (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=25, value="Консультування докторантів здобувачів (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=26, value="Керівництво ФПК (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=27, value="Робота приймальної комісії (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=28, value="Інше (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=29, value="Всього (перший семестр)")

    sheet_StudyLoadPerson.cell(row=1, column=30, value="Ставка (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=31, value="Знак сумісності (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=32, value="Лекції (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=33, value="Практичні (семінарські) заняття (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=34, value="Лабораторні роботи (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=35, value="Екзамени (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=36, value="Консультації перед екзаменами (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=37, value="Заліки (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=38, value="Кваліфікаційні робота (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=39, value="Атестаційні екзамени (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=40, value="Виробнича практика (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=41, value="Навчальна практика (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=42, value="Поточні консультації (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=43, value="Індивідуальні (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=44, value="Курсові роботи (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=45, value="Аспірантські екзамени (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=46, value="Керівництво аспірантами (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=47, value="Консультування докторантів здобувачів (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=48, value="Керівництво ФПК (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=49, value="Робота приймальної комісії (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=50, value="Інше (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=51, value="Всього (другий семестр)")

    sheet_StudyLoadPerson.cell(row=1, column=52, value="Ставка (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=53, value="Знак сумісності (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=54, value="Лекції (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=55, value="Практичні (семінарські) заняття (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=56, value="Лабораторні роботи (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=57, value="Екзамени (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=58, value="Консультації перед екзаменами (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=59, value="Заліки (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=60, value="Кваліфікаційні робота (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=61, value="Атестаційні екзамени (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=62, value="Виробнича практика (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=63, value="Навчальна практика (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=64, value="Поточні консультації (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=65, value="Індивідуальні (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=66, value="Курсові роботи (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=67, value="Аспірантські екзамени (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=68, value="Керівництво аспірантами (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=69, value="Консультування докторантів здобувачів (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=70, value="Керівництво ФПК (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=71, value="Робота приймальної комісії (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=72, value="Інше (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=73, value="Всього (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=74, value="Розподіл ставок навчального навантаження (рік)")


def print_db_with_vacancy(sheet_StudyLoadPerson):
    sheet_StudyLoadPerson.cell(row=1, column=1, value="Назва кафедри")
    sheet_StudyLoadPerson.cell(row=1, column=2, value="Назва")
    sheet_StudyLoadPerson.cell(row=1, column=3, value="Номер")
    sheet_StudyLoadPerson.cell(row=1, column=4, value="Посада")
    sheet_StudyLoadPerson.cell(row=1, column=5, value="Вчене звання, вчена ступінь")
    sheet_StudyLoadPerson.cell(row=1, column=6, value="Роки")

    sheet_StudyLoadPerson.cell(row=1, column=7, value="Ставка (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=8, value="Знак сумісності (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=9, value="Лекції (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=10, value="Практичні (семінарські) заняття (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=11, value="Лабораторні роботи (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=12, value="Екзамени (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=13, value="Консультації перед екзаменами (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=14, value="Заліки (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=15, value="Кваліфікаційні робота (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=16, value="Атестаційні екзамени (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=17, value="Виробнича практика (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=18, value="Навчальна практика (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=19, value="Поточні консультації (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=20, value="Індивідуальні (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=21, value="Курсові роботи (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=22, value="Аспірантські екзамени (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=23, value="Керівництво аспірантами (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=24, value="Консультування докторантів здобувачів (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=25, value="Керівництво ФПК (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=26, value="Робота приймальної комісії (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=27, value="Інше (перший семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=28, value="Всього (перший семестр)")

    sheet_StudyLoadPerson.cell(row=1, column=29, value="Ставка (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=30, value="Знак сумісності (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=31, value="Лекції (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=32, value="Практичні (семінарські) заняття (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=33, value="Лабораторні роботи (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=34, value="Екзамени (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=35, value="Консультації перед екзаменами (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=36, value="Заліки (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=37, value="Кваліфікаційні робота (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=38, value="Атестаційні екзамени (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=39, value="Виробнича практика (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=40, value="Навчальна практика (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=41, value="Поточні консультації (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=42, value="Індивідуальні (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=43, value="Курсові роботи (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=44, value="Аспірантські екзамени (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=45, value="Керівництво аспірантами (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=46, value="Консультування докторантів здобувачів (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=47, value="Керівництво ФПК (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=48, value="Робота приймальної комісії (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=49, value="Інше (другий семестр)")
    sheet_StudyLoadPerson.cell(row=1, column=50, value="Всього (другий семестр)")

    sheet_StudyLoadPerson.cell(row=1, column=51, value="Ставка (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=52, value="Знак сумісності (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=53, value="Лекції (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=54, value="Практичні (семінарські) заняття (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=55, value="Лабораторні роботи (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=56, value="Екзамени (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=57, value="Консультації перед екзаменами (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=58, value="Заліки (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=59, value="Кваліфікаційні робота (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=60, value="Атестаційні екзамени (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=61, value="Виробнича практика (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=62, value="Навчальна практика (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=63, value="Поточні консультації (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=64, value="Індивідуальні (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=65, value="Курсові роботи (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=66, value="Аспірантські екзамени (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=67, value="Керівництво аспірантами (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=68, value="Консультування докторантів здобувачів (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=69, value="Керівництво ФПК (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=70, value="Робота приймальної комісії (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=71, value="Інше (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=72, value="Всього (рік)")
    sheet_StudyLoadPerson.cell(row=1, column=73, value="Розподіл ставок навчального навантаження (рік)")


def show_button_click(root,name_db):
    file = filedialog.askopenfilename(
        title="Оберіть Excel файл",
        filetypes=[("Excel Files", "*.xlsx")]
    )

    if file == "":
        messagebox.showinfo("Помилка", "Файл не обрано або має валідну помилку в назві!")
        return

    loading = loading_window(root)


    conn = sqlite3.connect(name_db)

    cursor = conn.cursor()

    job_info = cursor.execute("SELECT * FROM РОБОЧА_ІНФОРМАЦІЯ").fetchall()
    person_info = cursor.execute("SELECT * FROM ЛЮДСЬКА_ІНФОРМАЦІЯ").fetchall()
    vacancy_info = cursor.execute("SELECT * FROM ІНФОРМАЦІЯ_ВАКАНСІЯ").fetchall()
    first_semester_info = cursor.execute("SELECT * FROM ПЕРШИЙ_СЕМЕСТР").fetchall()
    second_semester_info = cursor.execute("SELECT * FROM ДРУГИЙ_СЕМЕСТР").fetchall()
    academic_year_info = cursor.execute("SELECT * FROM АКАДЕМІЧНИЙ_РІК").fetchall()
    code_and_year_info = cursor.execute("SELECT * FROM КОД_РІК").fetchall()

    conn.close()

    wb = openpyxl.load_workbook(file)

    if 'Person' in wb.sheetnames:
        del wb['Person']
    if 'Vacancy' in wb.sheetnames:
        del wb['Vacancy']

    person_sheet = wb.create_sheet('Person')
    vacancy_sheet = wb.create_sheet('Vacancy')

    print_db_with_person(person_sheet)
    print_db_with_vacancy(vacancy_sheet)

    index_person = 2
    index_vacancy = 2

    try:
        for i in range(0,len(person_info)):
            if person_info[i][1] != "":
                person_sheet.cell(row=index_person, column=1, value=code_and_year_info[i][1])
                person_sheet.cell(row=index_person, column=2, value=person_info[i][1])
                person_sheet.cell(row=index_person, column=3, value=person_info[i][2])
                person_sheet.cell(row=index_person, column=4, value=person_info[i][3])
                person_sheet.cell(row=index_person, column=5, value=job_info[i][1])
                person_sheet.cell(row=index_person, column=6, value=job_info[i][2])
                person_sheet.cell(row=index_person, column=7, value=code_and_year_info[i][2])
                j = 1
                k = 8
                while j < len(first_semester_info[i]) - 1:
                    person_sheet.cell(row=index_person, column=k, value=first_semester_info[i][j])
                    person_sheet.cell(row=index_person, column=k + len(first_semester_info[i]) - 2,
                                      value=second_semester_info[i][j])
                    person_sheet.cell(row=index_person,
                                      column=k + len(first_semester_info[i]) + len(second_semester_info[i]) - 4,
                                      value=academic_year_info[i][j])
                    k += 1
                    j += 1
                person_sheet.cell(row=index_person, column=k + len(first_semester_info[i]) + len(
                    second_semester_info[i]) - 4, value=academic_year_info[i][j])
                index_person += 1
            else:
                vacancy_sheet.cell(row=index_vacancy, column=1, value=code_and_year_info[i][1])
                vacancy_sheet.cell(row=index_vacancy, column=2, value=vacancy_info[i][1])
                vacancy_sheet.cell(row=index_vacancy, column=3, value=vacancy_info[i][2])
                vacancy_sheet.cell(row=index_vacancy, column=4, value=job_info[i][1])
                vacancy_sheet.cell(row=index_vacancy, column=5, value=job_info[i][2])
                vacancy_sheet.cell(row=index_vacancy, column=6, value=code_and_year_info[i][2])
                j = 1
                k = 7
                while j < len(first_semester_info[i]) - 1:
                    vacancy_sheet.cell(row=index_vacancy, column=k, value=first_semester_info[i][j])
                    vacancy_sheet.cell(row=index_vacancy, column=k + len(first_semester_info[i]) - 2,
                                       value=second_semester_info[i][j])
                    vacancy_sheet.cell(row=index_vacancy,
                                       column=k + len(first_semester_info[i]) + len(second_semester_info[i]) - 4,
                                       value=academic_year_info[i][j])
                    k += 1
                    j += 1
                vacancy_sheet.cell(row=index_vacancy,
                                   column=k + len(first_semester_info[i]) + len(second_semester_info[i]) - 4,
                                   value=academic_year_info[i][j])
                index_vacancy += 1

    except Exception:
        j=1

    try:
        wb.save(file)
    except Exception:
        loading.destroy()
        messagebox.showinfo("Помилка", "Файл відкритий!\n\nЗакрийте файл та повторіть спробу!")
        return


    loading.destroy()

